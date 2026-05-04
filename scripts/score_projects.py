"""
score_projects.py — VEIR GTM Ranking Model
===========================================
Reads:   output/project_joined_v2.xlsx
Writes:  output/project_scored.xlsx

Two scoring models — clean separation by product line (per Priyanshi email):
  Indoor model (primary)       — uses ONLY Inside-rated attributes
  Underground model (secondary) — uses ONLY Outside-rated attributes
  Attributes rated on BOTH sides appear in both models with their respective weight.
  Attributes with no rating on a given side are fully excluded from that model.

Scoring pipeline:
  1. Convert customer_type_label → numeric
  2. Compress labor_cost_burden_score from [2,9] → [4,7] (range was too extreme)
  3. Normalize every attribute to 0–10 via percentile rank (robust to outliers)
  4. Invert time_to_power_months (lower months = more urgent = higher score)
  5. Compute weighted average for each model (NaN-safe: skip missing attributes)
  6. Rank all 554 projects per model
  7. Calibration check: report where the 7 VEIR-highlighted projects land
"""

import os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT      = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
IN_EXCEL  = os.path.join(ROOT, 'output', 'project_joined_v2.xlsx')
OUT_EXCEL = os.path.join(ROOT, 'output', 'project_scored.xlsx')

# ── Customer type → numeric (binary: hyperscaler vs non-hyperscaler) ───────────
# VEIR's stated preference is non-hyperscaler. No evidence for sub-ranking within
# non-hyperscaler categories, so all non-hyperscalers get the same score.
CUSTOMER_TYPE_SCORES = {
    'enterprise':  8,
    'neocloud':    8,
    'colo':        8,
    'oem':         8,
    'hyperscaler': 2,
}

# ── Attribute definitions ──────────────────────────────────────────────────────
# (column, inside_weight, outside_weight, direction, display_label)
#
# Weights directly from VEIR Review sheet:
#   High = 1.0 | Med = 0.5 | Low = 0.2 | High/Med = 0.75 | Med/Low = 0.35 | — = 0.0
#
# inside_weight = 0.0  → attribute excluded from Indoor model entirely
# outside_weight = 0.0 → attribute excluded from Underground model entirely
# Both > 0             → appears in both models with respective side's weight
#
# Direction:
#   'higher_better' — higher raw value → higher score for VEIR
#   'lower_better'  — lower raw value  → higher score for VEIR (inverted)

ATTRIBUTES = [
    # col                                  inside  outside  direction           label
    ('time_to_power_months',                1.0,    1.0,    'lower_better',    'Time to Power (months)'),
    ('whitespace_power_density_w_sqft',     1.0,    0.0,    'higher_better',   'Whitespace Power Density (W/sqft)'),
    ('campus_power_density_w_sqft',         0.0,    1.0,    'higher_better',   'Campus Power Density (W/sqft)'),
    ('cable_multiplicity_burden_score',     0.0,    0.2,    'higher_better',   'Cable Multiplicity Burden'),
    ('whitespace_congestion_burden_score',  0.2,    0.0,    'higher_better',   'Whitespace Congestion Burden'),
    ('resiliency_topology_complexity_score',0.0,    0.2,    'higher_better',   'Resiliency Topology Complexity'),
    ('high_density_load_share_score',       0.5,    0.5,    'higher_better',   'High-Density Load Share'),
    ('foak_commercial_access_score',        0.5,    0.5,    'higher_better',   'FOAK Commercial Access'),
    ('customer_type_score',                 0.5,    0.5,    'higher_better',   'Customer Type'),
    ('permitting_burden_index',             0.5,    1.0,    'higher_better',   'Permitting Burden'),
    ('hv_mv_readiness_gap_score',           0.0,    0.5,    'higher_better',   'HV/MV Readiness Gap'),
    ('land_cost_proxy_score',               0.5,    0.75,   'higher_better',   'Land Cost Proxy'),
    ('buildable_land_constraint_score',     0.35,   0.75,   'higher_better',   'Buildable Land Constraint'),
    ('campus_size_sqft',                    0.5,    0.5,    'higher_better',   'Campus Size (sqft)'),
    ('labor_cost_burden_score',             1.0,    1.0,    'higher_better',   'Labor Cost Burden'),
]

# Indoor effective weight  = inside_weight  (outside fully excluded)
# Underground effective weight = outside_weight (inside fully excluded)
for i, (col, ins, out, direction, label) in enumerate(ATTRIBUTES):
    ATTRIBUTES[i] = (col, ins, out, direction, label, ins, out)

# ── 7 VEIR-highlighted projects (calibration anchors) ─────────────────────────
HIGHLIGHTED = [
    'SULPHER SPRINGS MATRIX DATA CENTER CAMPUS EXPANSION PH III BUILDING 28',
    'SPRINGFIELD DATA CENTER EXPANSION BUILDING 5',
    'YORKVILLE (PROJECT CARDINAL) DATA CENTER CAMPUS EXPANSION BUILDING 8',
    'MAXWELL CALDWELL VALLEY DATA CENTER EXPANSION BUILDING 10',
    'TAYLORSVILLE DATA CENTER CAMPUS EXPANSION BUILDING 7',
    'EVANSTON (ASPEN MOUNTAIN) DATA CENTER EXPANSION PHASE II BUILDING 4',
    'AMARILLO DATA CENTER CAMPUS (PROJECT MATADOR) EXPANSION BUILDING 13',
]

# ══════════════════════════════════════════════════════════════════════════════
# LOAD
# ══════════════════════════════════════════════════════════════════════════════
print(f"Loading: {IN_EXCEL}")
df = pd.read_excel(IN_EXCEL, sheet_name='project_joined_v2')
print(f"  {len(df)} rows, {len(df.columns)} columns\n")

# ── Customer type → numeric ────────────────────────────────────────────────────
df['customer_type_score'] = (
    df['customer_type_label'].str.lower().str.strip()
    .map(CUSTOMER_TYPE_SCORES)
)
unmapped = df['customer_type_score'].isna().sum()
if unmapped:
    print(f"  WARNING: {unmapped} rows have unmapped customer_type_label")
    print(df[df['customer_type_score'].isna()]['customer_type_label'].value_counts())

# ── Compress labor_cost_burden_score from [2, 9] → [4, 7] ─────────────────────
# The LLM-scored range was too extreme: high-union states (CA/IL/PA) got 8-9,
# right-to-work states (TX/GA/WY) got 2-3, a 7-point gap.
# VEIR's labor savings from fewer cable pulls is real in all states — the
# marginal benefit scales more moderately than the raw scores implied.
# Linear remap: new = 4 + (old - 2) / 7 × 3  →  [2,9] maps to [4, 7]
_labor_min, _labor_max = 2.0, 9.0
_new_min,   _new_max   = 4.0, 7.0
df['labor_cost_burden_score'] = (
    _new_min + (df['labor_cost_burden_score'] - _labor_min)
    / (_labor_max - _labor_min) * (_new_max - _new_min)
).clip(_new_min, _new_max)
print(f"\nLabor score compressed [2,9]→[4,7]:  "
      f"min={df['labor_cost_burden_score'].min():.1f}  "
      f"max={df['labor_cost_burden_score'].max():.1f}  "
      f"mean={df['labor_cost_burden_score'].mean():.2f}\n")

# ══════════════════════════════════════════════════════════════════════════════
# NORMALIZE each attribute to 0–10 via percentile rank
# ══════════════════════════════════════════════════════════════════════════════
print("Normalizing attributes (percentile rank → 0–10)...")
print(f"  {'Attribute':40s}  {'Indoor w':>8}  {'Undergnd w':>10}  Direction")
print(f"  {'-'*40}  {'-'*8}  {'-'*10}  ---------")

NORM_COLS = {}   # col → normalized column name

for col, ins, out, direction, label, indoor_eff, underground_eff in ATTRIBUTES:
    if col not in df.columns:
        print(f"  SKIP (missing): {col}")
        continue

    norm_col = f'_norm_{col}'
    df[norm_col] = df[col].rank(pct=True, na_option='keep') * 10.0

    if direction == 'lower_better':
        df[norm_col] = 10.0 - df[norm_col]

    NORM_COLS[col] = norm_col
    in_tag  = f"{indoor_eff:.1f}" if indoor_eff > 0 else "excl."
    out_tag = f"{underground_eff:.2f}" if underground_eff > 0 else "excl."
    print(f"  {label:40s}  {in_tag:>8}  {out_tag:>10}  [{direction}]")

# ══════════════════════════════════════════════════════════════════════════════
# SCORE — weighted average (NaN-safe)
# ══════════════════════════════════════════════════════════════════════════════
print("\nComputing Indoor and Underground scores...")

indoor_scores     = np.zeros(len(df))
underground_scores = np.zeros(len(df))
indoor_weights_sum     = np.zeros(len(df))
underground_weights_sum = np.zeros(len(df))

for col, ins, out, direction, label, indoor_eff, underground_eff in ATTRIBUTES:
    if col not in NORM_COLS:
        continue
    norm_col = NORM_COLS[col]
    valid    = df[norm_col].notna()

    indoor_scores[valid]     += df.loc[valid, norm_col] * indoor_eff
    underground_scores[valid] += df.loc[valid, norm_col] * underground_eff
    indoor_weights_sum[valid]     += indoor_eff
    underground_weights_sum[valid] += underground_eff

# Normalize to 0–10 weighted average
df['indoor_score']     = np.where(
    indoor_weights_sum > 0,
    indoor_scores / indoor_weights_sum * 10 / 10,   # already 0–10
    np.nan
)
df['underground_score'] = np.where(
    underground_weights_sum > 0,
    underground_scores / underground_weights_sum * 10 / 10,
    np.nan
)

# Rank (1 = best)
df['indoor_rank']     = df['indoor_score'].rank(ascending=False, method='min').astype(int)
df['underground_rank'] = df['underground_score'].rank(ascending=False, method='min').astype(int)

# Highlighted flag
df['veir_highlighted'] = df['Project Name'].isin(HIGHLIGHTED)

print(f"  Indoor score:     mean={df['indoor_score'].mean():.2f}  "
      f"min={df['indoor_score'].min():.2f}  max={df['indoor_score'].max():.2f}")
print(f"  Underground score: mean={df['underground_score'].mean():.2f}  "
      f"min={df['underground_score'].min():.2f}  max={df['underground_score'].max():.2f}")

# ══════════════════════════════════════════════════════════════════════════════
# CALIBRATION CHECK — where do the 7 highlighted projects rank?
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== CALIBRATION CHECK (7 VEIR-highlighted projects) ===")
h = df[df['veir_highlighted']][
    ['Project Name', 'State', 'customer_type_label',
     'indoor_score', 'indoor_rank', 'underground_score', 'underground_rank']
].sort_values('indoor_rank')

for _, r in h.iterrows():
    indoor_pct     = (1 - r['indoor_rank'] / len(df)) * 100
    underground_pct = (1 - r['underground_rank'] / len(df)) * 100
    print(f"  #{r['indoor_rank']:3d} ({indoor_pct:4.1f}%ile) indoor  |  "
          f"#{r['underground_rank']:3d} ({underground_pct:4.1f}%ile) underground  |  "
          f"{r['Project Name'][:48]}  [{r['customer_type_label']}]")

h_indoor_ranks = h['indoor_rank'].tolist()
print(f"\n  Indoor:     avg rank = {np.mean(h_indoor_ranks):.0f} / {len(df)}  "
      f"(top {np.mean(h_indoor_ranks)/len(df)*100:.1f}%)")
print(f"  In top 50:  {sum(r <= 50 for r in h_indoor_ranks)} / 7 highlighted projects")
print(f"  In top 100: {sum(r <= 100 for r in h_indoor_ranks)} / 7 highlighted projects")

# ══════════════════════════════════════════════════════════════════════════════
# BUILD OUTPUT EXCEL
# ══════════════════════════════════════════════════════════════════════════════
print("\nBuilding output Excel...")

# Key columns for the ranked output
KEY_COLS = [
    'indoor_rank', 'underground_rank',
    'indoor_score', 'underground_score',
    'veir_highlighted',
    'Project Name', 'Owner Name', 'City', 'State',
    'customer_type_label', 'Project Capacity (Numeric)',
    'time_to_power_months',
    'whitespace_power_density_w_sqft',
    'campus_power_density_w_sqft',
    'campus_size_sqft',
    'high_density_load_share_score',
    'foak_commercial_access_score',
    'labor_cost_burden_score',
    'permitting_burden_index',
    'hv_mv_readiness_gap_score',
    'land_cost_proxy_score',
    'buildable_land_constraint_score',
]

ranked = df[KEY_COLS].copy()
ranked_indoor = ranked.sort_values('indoor_rank').reset_index(drop=True)
ranked_underground = ranked.sort_values('underground_rank').reset_index(drop=True)

# Weight reference table
W_LABEL = {1.0:'High', 0.75:'High/Med', 0.5:'Med', 0.35:'Med/Low', 0.2:'Low', 0.0:'—'}
weight_rows = []
for col, ins, out, direction, label, indoor_eff, underground_eff in ATTRIBUTES:
    weight_rows.append({
        'Attribute': label,
        'Column': col,
        'Inside Rating (Review)':  W_LABEL.get(ins, str(ins)),
        'Outside Rating (Review)': W_LABEL.get(out, str(out)),
        'Direction': direction,
        'Indoor Weight (Inside only)':      round(indoor_eff, 2) if indoor_eff > 0 else 'excluded',
        'Underground Weight (Outside only)': round(underground_eff, 2) if underground_eff > 0 else 'excluded',
    })
weights_df = pd.DataFrame(weight_rows)

# Calibration table
calib_rows = []
for _, r in h.sort_values('indoor_rank').iterrows():
    calib_rows.append({
        'Project Name': r['Project Name'],
        'State': r['State'],
        'Customer Type': r['customer_type_label'],
        'Indoor Score': round(r['indoor_score'], 2),
        'Indoor Rank': int(r['indoor_rank']),
        'Indoor Percentile': f"Top {(1 - r['indoor_rank']/len(df))*100:.1f}%",
        'Underground Score': round(r['underground_score'], 2),
        'Underground Rank': int(r['underground_rank']),
        'Underground Percentile': f"Top {(1 - r['underground_rank']/len(df))*100:.1f}%",
    })
calib_df = pd.DataFrame(calib_rows)

# Write Excel
with pd.ExcelWriter(OUT_EXCEL, engine='openpyxl') as writer:
    # Sheet 1: Indoor ranking (all 554)
    ranked_indoor.to_excel(writer, sheet_name='Indoor Ranking', index=False)
    # Sheet 2: Underground ranking (all 554)
    ranked_underground.to_excel(writer, sheet_name='Underground Ranking', index=False)
    # Sheet 3: Top 20 Indoor
    ranked_indoor.head(20).to_excel(writer, sheet_name='Top 20 Indoor', index=False)
    # Sheet 4: Top 20 Underground
    ranked_underground.head(20).to_excel(writer, sheet_name='Top 20 Underground', index=False)
    # Sheet 5: Top 20 Unique Campuses — Indoor
    # Per campus (Umbrella Project ID), keep only the best-ranked phase.
    # Projects with no Umbrella ID are treated as standalone campuses.
    df_full = pd.read_excel(IN_EXCEL, sheet_name='project_joined_v2',
                            usecols=['Project Name', 'Umbrella Project ID', 'Umbrella Project Name'])
    ranked_indoor_full = ranked_indoor.merge(
        df_full[['Project Name', 'Umbrella Project ID', 'Umbrella Project Name']],
        on='Project Name', how='left'
    )
    # Fill missing Umbrella Project ID with the Project Name itself (standalone)
    ranked_indoor_full['_campus_key'] = ranked_indoor_full['Umbrella Project ID'].fillna(
        ranked_indoor_full['Project Name']
    )
    top20_unique_indoor = (
        ranked_indoor_full
        .sort_values('indoor_rank')
        .drop_duplicates(subset='_campus_key', keep='first')
        .head(20)
        .drop(columns=['_campus_key', 'Umbrella Project ID', 'Umbrella Project Name'])
        .reset_index(drop=True)
    )
    top20_unique_indoor.to_excel(writer, sheet_name='Top 20 Unique Campuses', index=False)

    # Sheet 6: Weight table
    weights_df.to_excel(writer, sheet_name='Weight Reference', index=False)
    # Sheet 7: Calibration
    calib_df.to_excel(writer, sheet_name='Calibration Check', index=False)

# ── Apply yellow highlighting to VEIR-selected projects ────────────────────────
wb = openpyxl.load_workbook(OUT_EXCEL)
YELLOW = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
HEADER_FILL = PatternFill(start_color='FF1F3864', end_color='FF1F3864', fill_type='solid')
HEADER_FONT = Font(color='FFFFFFFF', bold=True)

for sheet_name in ['Indoor Ranking', 'Underground Ranking', 'Top 20 Indoor', 'Top 20 Underground', 'Top 20 Unique Campuses']:
    ws = wb[sheet_name]
    # Style header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    # Find 'Project Name' column and 'veir_highlighted' column
    header = {cell.value: cell.column for cell in ws[1]}
    proj_col    = header.get('Project Name')
    highlight_col = header.get('veir_highlighted')
    score_col   = header.get('indoor_score') or header.get('underground_score')
    # Yellow-fill highlighted rows
    for row in ws.iter_rows(min_row=2):
        proj_name = row[proj_col - 1].value if proj_col else None
        if proj_name in HIGHLIGHTED:
            for cell in row:
                cell.fill = YELLOW
    # Auto-width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)

# Style Weight Reference sheet
ws = wb['Weight Reference']
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

# Style Calibration sheet
ws = wb['Calibration Check']
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.fill = YELLOW

wb.save(OUT_EXCEL)
print(f"\n=== Done ===")
print(f"  Output: {OUT_EXCEL}")
print(f"  Sheets: Indoor Ranking | Underground Ranking | Top 20 Indoor | Top 20 Underground | Top 20 Unique Campuses | Weight Reference | Calibration Check")
